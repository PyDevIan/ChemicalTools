import tkinter as tk
from tkinter import font
import sys
from tkinter import messagebox
import psutil
import openpyxl
from datetime import datetime, timezone

# Create object
root =tk.Tk()
  
# Adjust size
root.geometry( "320x300" )
root.title("Brine Calculator")
custom_font = font.Font(size=16) 

# Συνταγές Άλμεων.(Σε Λίτρα και κιλά)
Brine_ISP_Tank = {"Be":9,"Lactic Acid": 600 ,"Acetico":100}
Brine_ISP_Drum = {"Be": 4682 , "Lactic Acid":240  }
Brine_KAL = {"Be": 4682 , "Lactic Acid":100 }
Brine_FYS_Tank = {"Be":6,"Acetico":1900}
Brine_FYS_Drum = {"Be":5,"Acetico":780}
Brine_Pepper = {"Be":18 , "Benzoic_Acid":50 , "Calcium Lactate":200 ,"Citric Acid": 300}


def write(x ,y ,z,w):
        file_path = 'TEST.xlsx'  # Replace with the actual file path
        book = openpyxl.load_workbook(file_path)
        sheet_name = w # Replace with the actual sheet name
        row_index = int(x)  # Replace with the desired row index
        column_name = str(y)  # Replace with the desired column name
        new_value = z # Replace with the new value
        
        
        sheet = book[sheet_name]
        column_index = openpyxl.utils.column_index_from_string(column_name)
        column_letter = openpyxl.utils.get_column_letter(column_index)
        cell_address = f"{column_letter}{row_index}"
        cell = sheet[cell_address]
        cell.value = new_value   
        book.save('TEST.xlsx')
        
        
        # Save the changes
        

def calculate():
    brine_level_start = float(entry1.get())
    Be_first= float(entry2.get())
    Be_final = float(entry3.get())
    Tank= int(entry4.get())
    file_path = 'TEST.xlsx'  # Replace with the actual file path
    book = openpyxl.load_workbook(file_path)
    TANKHIGHT=273,68
    TANKVOLUME=52
    brine_start=brine_level_start*0.19
    amount_manuf=TANKVOLUME- brine_start
    Salt_start= (brine_start*1000*Be_first)/100
    Salt_end= (TANKVOLUME*1000*Be_final)/100
    Saltwater_kg=(Salt_end-Salt_start)*100/20
    now = datetime.now(timezone.utc)
    formatted_time = now.strftime("%d-%m-%y")
    Lot = now.strftime("%y%m%d")
    
    Brine = clicked.get()
    if Brine == "Ισπανική Δεξαμενισμού" and Tank== 1 :
        Lactic_Acid =round((amount_manuf*Brine_ISP_Tank['Lactic Acid']/TANKVOLUME) ,1)
        Lactic_Acid_kg=round((Lactic_Acid*1.18),1)
        Acetic_Acid=round((amount_manuf*Brine_ISP_Tank['Acetico']/TANKVOLUME),1)
        Acetic_Acid_Kg=round((Acetic_Acid*1.06),1)
        write(15,"E",brine_start,'ΑΛΜΗΣ 3')
        write(16,'E',amount_manuf,'ΑΛΜΗΣ 3')
        write(10,'C',Saltwater_kg,'ΑΛΜΗΣ 3')
        write(10,'E',Lactic_Acid_kg,'ΑΛΜΗΣ 3')
        write(10,'G',Acetic_Acid_Kg,'ΑΛΜΗΣ 3')
        write(2,'H',formatted_time,'ΑΛΜΗΣ 3')
        write(2,'E',Lot,'ΑΛΜΗΣ 3')
        write(3,'H',Tank,'ΑΛΜΗΣ 3')
        
        return label5.config(text=f'Παρασκευ. Ποσότητα (m3): {amount_manuf} \n Be τελικό : {Be_final} \n Καταναλώθηκαν: \n Αλατόνερο 20%: {Saltwater_kg} Kg'+
                         f' \n Γαλακτικό οξύ: {Lactic_Acid} L / {Lactic_Acid_kg} Kg \n Οξικό οξύ: {Acetic_Acid} L \ {Acetic_Acid_Kg} kg ')  
      
    elif Brine == "Ισπανική Βαρέλια":
        Lactic_Acid =round((amount_manuf*Brine_ISP_Drum['Lactic Acid']/TANKVOLUME) ,1)
        Lactic_Acid_kg=round((Lactic_Acid*1.18),1)
        
        return label5.config(text=f'Παρασκευ. Ποσότητα (m3): {amount_manuf} \n Be τελικό : {Be_final} \n Καταναλώθηκαν: \n Αλατόνερο 20%: {Saltwater_kg} Kg'+
                         f' \n Γαλακτικό οξύ: {Lactic_Acid} L / {Lactic_Acid_kg} Kg')
        
    elif Brine == "Καλαμών" :
        Lactic_Acid =round((amount_manuf*Brine_KAL['Lactic Acid']/TANKVOLUME) ,1)
        Lactic_Acid_kg=round((Lactic_Acid*1.18),1)
        return label5.config(text=f'Παρασκευ. Ποσότητα (m3): {amount_manuf} \n Be τελικό : {Be_final} \n Καταναλώθηκαν: \n Αλατόνερο 20%: {Saltwater_kg} Kg'+
                         f' \n Γαλακτικό οξύ: {Lactic_Acid} L / {Lactic_Acid_kg} Kg')
    
    elif Brine == "Φυσική Δεξαμενισμού":
        Acetic_Acid=round((amount_manuf*Brine_FYS_Tank['Acetico']/TANKVOLUME),1)
        Acetic_Acid_Kg=round((Acetic_Acid*1.06),1)
        return label5.config(text=f'Παρασκευ. Ποσότητα (m3): {amount_manuf} \n Be τελικό : {Be_final} \n Καταναλώθηκαν: \n Αλατόνερο 20%: {Saltwater_kg} Kg'+
                         f' \n Οξικό οξύ: {Acetic_Acid} L \ {Acetic_Acid_Kg} kg ')  
    
    elif Brine == "Φυσική Βαρέλια":
        Acetic_Acid=round((amount_manuf*Brine_FYS_Drum['Acetico']/TANKVOLUME),1)
        Acetic_Acid_Kg=round((Acetic_Acid*1.06),1)
        return label5.config(text=f'Παρασκευ. Ποσότητα (m3): {amount_manuf} \n Be τελικό : {Be_final} \n Καταναλώθηκαν: \n Αλατόνερο 20%: {Saltwater_kg} Kg'+
                         f' \n Οξικό οξύ: {Acetic_Acid} L \ {Acetic_Acid_Kg} kg ')  
    
    elif Brine == "Καρδούλα":
        Benzoic_Acid=round((amount_manuf*Brine_Pepper['Benzoic_Acid']/TANKVOLUME),1)
        Calcium_Lactate=round((amount_manuf*Brine_Pepper['Calcium Lactate']/TANKVOLUME),1)
        Citric_Acid=round((amount_manuf*Brine_Pepper['Citric Acid']/TANKVOLUME),1)
        return label5.config(text=f'Παρασκευ. Ποσότητα (m3): {amount_manuf} \n Be τελικό : {Be_final} \n Καταναλώθηκαν: \n Αλατόνερο 20%: {Saltwater_kg} Kg'+
                         f' \n Βενζοικό Οξύ: {Benzoic_Acid} Kg  \n Γαλακτικό Ασβέστιο: {Calcium_Lactate} Kg \n Κιτρικό Οξύ: {Citric_Acid} Kg')
        
    else:
      return label5.config(text="Not a valid choice",font=custom_font, fg='red')


label1 = tk.Label(root, text="Ύψος στάθμης αρχικό (cm):")
label1.pack()
entry1 = tk.Entry(root)
entry1.pack()

label2 = tk.Label(root, text="Βe αρχικό:")
label2.pack()
entry2 = tk.Entry(root)
entry2.pack()

label3 = tk.Label(root, text="Βe τελικό:")
label3.pack()
entry3 = tk.Entry(root)
entry3.pack()

label4 = tk.Label(root, text="Δεξαμενή:")
label4.pack()
entry4 = tk.Entry(root)
entry4.pack()

Brines_List=["Ισπανική Δεξαμενισμού", "Ισπανική Βαρέλια", "Καλαμών", "Φυσική Δεξαμενισμού",
             "Φυσική Βαρέλια", "Καρδούλα"]

clicked = tk.StringVar()
  
# initial menu text
clicked.set( "Choose Brine Type" )
  
# Create Dropdown menu
drop = tk.OptionMenu( root , clicked , *Brines_List ,)
drop.pack()
  
# Create button, it will change label text
button =tk.Button( root , text = "Calculate & Export" , command = calculate ).pack()

  
 
label5=tk.Label(root,text="")
label5.pack()

def main():
 
  root.mainloop()

def check_already_running(process_name):
    for process in psutil.process_iter(['name']):
        if process.info['name'] == process_name:
            messagebox.showerror('Error', 'Your Application is already running!')
            sys.exit(0)

if __name__ == "__main__":
    process_name = "BrineCalc"

    check_already_running(process_name)

    main()